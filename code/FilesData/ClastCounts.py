import openpyxl
import numpy as np
from matplotlib import pyplot as plt

# Input variables
file_name = '../../data/ClastCountCompilation_sample.xlsx'
wb = openpyxl.load_workbook(file_name)

def values_in_column(sheet, n, m=8):
  outlist = []
  cells = sheet.columns[n][m:]
  for cell in cells:
    value = cell.value
    if value:
      outlist.append(value)
    else:
      outlist.append(np.nan)
  outarray = np.array(outlist)
  return outarray

# Starting simple: just location and grain size
lats = []
lons = []
Ds   = []
lithologies = []
for sheet in wb:
  print sheet.title
  try:
    lat = -sheet.cell(coordinate='C1').value
    lon = -sheet.cell(coordinate='C2').value
  except:
    print "  >> NO LOCATION"
    lat = np.nan
    lon = np.nan
  D = values_in_column(sheet, 1)
  lith = values_in_column(sheet, 2)
  lith = lith[np.isnan(D) == False]
  D = D[np.isnan(D) == False]
  D = D.astype(float)
  print len(D)
  lats.append(lat)
  lons.append(lon)
  Ds.append(D)
  lithologies.append(lith)

# Arrange the grain size classes
phi = -np.arange(1,12)
Dmin = 2**-phi
Dmax = Dmin[1:]
Dmin = Dmin[:-1]
Dmean = (Dmin + Dmax)/2.

Dclassified = []
for D in Ds:
  Dclasses = []
  Dclasses.append(np.percentile(D, 50)) # D50
  Dclasses.append(np.percentile(D, 84)) # D84
  for i in range(len(Dmin)):
    Dclasses.append(np.sum( (D >= Dmin[i]) * (D < Dmax[i]) ))
  Dclassified.append(Dclasses)
  

# Lats and lons
ll = np.zeros((len(lats), 2))
ll[:,0] = lons
ll[:,1] = lats

# Create a GRASS-formatted vector ASCII

# Generate header
hdr_coor = ['lon', 'lat']
hdr_site = ['site']
hdr_D50  = ['D50']
hdr_D84  = ['D84']
hdr_D    = list(Dmean.astype(int).astype(str))
hdr = hdr_coor + hdr_site + hdr_D50 + hdr_D84 + hdr_D

# Site names
sites = np.expand_dims( wb.sheetnames, 2)

# Stack data and header
hdr_array = np.expand_dims(np.array(hdr), 2).transpose().astype('|S18')
alldata = np.hstack((ll, sites, np.array(Dclassified))).astype('|S18')
alldata = alldata[np.isnan(lats) == False] # Remove those without coordinates
outarray = np.vstack((hdr_array, alldata))

# Output
np.savetxt('ClastCounts.txt', outarray, fmt='%s', delimiter='|')


try:
  from grass import script as g
  # Send classified D to GRASS
  g.run_command('v.in.ascii', input='ClastCounts.txt', output='ClastCounts', \
  skip=1, columns='lon double precision, lat double precision,site varchar, \
  D50 double precision, D84 double precision, D3mm double precision, \
  D6mm double precision, D12mm double precision, D24mm double precision, \
  D48mm double precision, D96mm double precision, D192mm double precision, \
  D384mm double precision, D768mm double precision, D1536mm double precision', \
  overwrite=True)
except:
  print "GRASS GIS could not be initialized: if you want to import this file,"
  print "please run it within a GRASS GIS environment."
  print "You may also consider modifying this code to work with basic"
  print "GDAL/OGR to perform the calculation with less software overhead."
